import openpyxl


def getRowCount(path,sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    totalrows = sheet.max_row
    return totalrows

def getColCount(path,sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    totalcols = sheet.max_column
    return totalcols


def getCellData(path,sheetName,rowNum,colNum):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    getData = sheet.cell(row=rowNum,column=colNum).value
    return getData


def setCellData(path,sheetName,rowNum,colNum,data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    getData = sheet.cell(row=rowNum, column=colNum).value = data
    workbook.save(path)

